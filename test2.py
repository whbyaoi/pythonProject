""" for yqq """
import traceback

import requests
import openpyxl

interval = 7


def get_json_content(url, waiting_time):
    """ return json content """
    import time
    import requests
    import json
    try:
        time.sleep(waiting_time)
        r = requests.get(url)
        json_content = json.loads(str(r.content, encoding="utf-8"))
        return json_content
    except:
        print('except occurred')
        print('url: {}'.format(url))
        print('next crawling interval: {}'.format(waiting_time + interval))
        return get_json_content(url, waiting_time + interval)


if __name__ == '__main__':
    book = openpyxl.load_workbook('20211008.xlsx')
    # print(book.sheetnames)
    sheet = book.worksheets[6]

    book2 = openpyxl.load_workbook('select.xlsx')
    # print(book2.sheetnames)
    sheet2 = book2.worksheets[6]
    count = 0
    for address in set([row[0].value for row in list(sheet.rows)[1:]]):
        if address:
            address_content = get_json_content('https://chain.api.btc.com/v3/address/' + address, interval)
            if 'last_tx' in address_content['data'] and address_content['data']['last_tx']:
                latest_tx_content = get_json_content('https://chain.api.btc.com/v3/tx/{}?verbose=1'.format(address_content['data']['last_tx']), interval)
                # print(sheet.title, address, latest_tx_content['data']['block_height'])
                if int(latest_tx_content['data']['block_height']) >= 585000:
                    print(sheet.title, address, latest_tx_content['data']['block_height'])
                    sheet2.append([address, latest_tx_content['data']['block_height']])
                    count += 1
                    if count % 100 == 0:
                        book2.save(str(sheet.title) + ".xlsx")
                else:
                    print("pass:", sheet.title, address, latest_tx_content['data']['block_height'])
            else:
                print("pass, reason: last_tx not exist, {}".format(address))
    book2.save(str(sheet.title) + ".xlsx")
